#!/usr/bin/env python3
# -*- coding: utf-8 -*-
try:
    import sys
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

import argparse, csv, json, sys
from pathlib import Path

def read_csv(p):
    with open(p, encoding='utf-8-sig', newline='') as f:
        return list(csv.reader(f))

def read_md(p):
    rows = []
    for ln in open(p, encoding='utf-8').read().splitlines():
        ln = ln.strip()
        if not ln.startswith('|') or not ln.endswith('|'): continue
        cells = [c.strip() for c in ln.strip('|').split('|')]
        if all(set(c) <= {'-',':',' '} for c in cells): continue
        rows.append(cells)
    return rows

def read_json(p):
    data = json.load(open(p, encoding='utf-8'))
    if isinstance(data, dict) and isinstance(data.get('data'), list): data = data['data']
    if not data: return []
    header = list(data[0].keys())
    return [header] + [[str(d.get(k,'')) for k in header] for d in data]

def to_md(rows):
    if not rows: return ''
    out = ['| ' + ' | '.join(rows[0]) + ' |', '|' + '|'.join(['---'] * len(rows[0])) + '|']
    for r in rows[1:]:
        out.append('| ' + ' | '.join(r) + ' |')
    return '\n'.join(out)

def to_csv(rows, p):
    with open(p, 'w', encoding='utf-8-sig', newline='') as f:
        csv.writer(f).writerows(rows)

def main():
    p = argparse.ArgumentParser()
    p.add_argument('convert'); p.add_argument('input'); p.add_argument('-o','--output', required=True)
    a = p.parse_args()
    src = Path(a.input)
    suffix = src.suffix.lower()
    try:
        if suffix == '.csv': rows = read_csv(src)
        elif suffix == '.md': rows = read_md(src)
        elif suffix == '.json': rows = read_json(src)
        elif suffix == '.xlsx':
            try:
                import openpyxl
            except ImportError:
                sys.stderr.write('[table_tool] xlsx 需 openpyxl: pip install "openpyxl>=3.0"\n'); sys.exit(1)
            wb = openpyxl.load_workbook(src, data_only=True)
            ws = wb[wb.sheetnames[0]]
            rows = [['' if x is None else str(x) for x in r] for r in ws.iter_rows(values_only=True)]
        else:
            sys.stderr.write('[table_tool] 不支持输入格式: ' + suffix + '\n'); sys.exit(1)
    except Exception as e:
        sys.stderr.write('[table_tool] 读取失败: ' + str(e) + '\n'); sys.exit(1)
    out = Path(a.output)
    if out.suffix.lower() == '.md':
        out.write_text(to_md(rows), encoding='utf-8')
    elif out.suffix.lower() == '.csv':
        to_csv(rows, out)
    elif out.suffix.lower() == '.html':
        out.write_text('<table border="1">\n' + '\n'.join('<tr>' + ''.join('<td>' + c + '</td>' for c in r) + '</tr>' for r in rows) + '\n</table>', encoding='utf-8')
    elif out.suffix.lower() == '.json':
        objs = [dict(zip(rows[0], r)) for r in rows[1:]]
        json.dump(objs, out.open('w', encoding='utf-8'), ensure_ascii=False, indent=2)
    else:
        sys.stderr.write('[table_tool] 输出需为 .md/.csv/.html/.json\n'); sys.exit(1)
    print('已转换 ' + str(src) + ' -> ' + str(out) + ' (' + str(max(len(rows)-1, 0)) + ' 行)')

main()
