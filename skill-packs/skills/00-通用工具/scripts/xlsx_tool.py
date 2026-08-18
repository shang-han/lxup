#!/usr/bin/env python3
# -*- coding: utf-8 -*-
try:
    import sys
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

import argparse, csv, json, sys
try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.stderr.write('[xlsx_tool] 缺少依赖 openpyxl,请先执行: pip install "openpyxl>=3.0"\n'); sys.exit(1)

THIN = Side(style='thin')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill('solid', fgColor='DDEBF7')
HEADER_FONT = Font(bold=True)

def rows_csv(path):
    with open(path, 'r', encoding='utf-8-sig', newline='') as f:
        return list(csv.reader(f))

def rows_json(path):
    data = json.load(open(path, encoding='utf-8'))
    if isinstance(data, dict):
        for k in ('data','items','rows'):
            if isinstance(data.get(k), list): data = data[k]; break
        else:
            sys.stderr.write('[xlsx_tool] JSON 需为数组或带 data/items/rows 键\n'); sys.exit(1)
    if not data or not isinstance(data[0], dict):
        sys.stderr.write('[xlsx_tool] JSON 需为对象数组(键名即列名)\n'); sys.exit(1)
    header = list(data[0].keys())
    return [header] + [[str(d.get(k,'')) for k in header] for d in data]

def append_row(ws, row):
    """写入一行;以 = 开头的字符串写为真实公式(如 =SUM(B2:B10))"""
    ws.append(row)
    r = ws.max_row
    for idx, val in enumerate(row, 1):
        if isinstance(val, str) and val.startswith('='):
            ws.cell(row=r, column=idx).value = val

def style_sheet(ws):
    """表头加粗+着色+居中、全表边框、按内容自动列宽(上限 40)、冻结首行"""
    for c in ws[1]:
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal='center')
    for row in ws.iter_rows():
        for c in row:
            c.border = BORDER
    widths = {}
    for row in ws.iter_rows(values_only=True):
        for i, v in enumerate(row, 1):
            if v is None:
                continue
            s = str(v)
            if s.startswith('='):  # 公式文本不参与列宽估算
                continue
            w = sum(2 if ord(ch) > 127 else 1 for ch in s)
            widths[i] = max(widths.get(i, 0), w)
    for i, w in widths.items():
        ws.column_dimensions[get_column_letter(i)].width = min(w + 2, 40)
    ws.freeze_panes = 'A2'

def cmd_create(a):
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    for src in (a.csv or []) + (a.json or []):
        name = src.rsplit('/',1)[-1].rsplit('\\',1)[-1].split('.')[0][:31] or 'Sheet'
        rows = rows_csv(src) if src in (a.csv or []) else rows_json(src)
        ws = wb.create_sheet(title=name)
        for r in rows:
            append_row(ws, r)
        style_sheet(ws)
    if not wb.sheetnames:
        sys.stderr.write('[xlsx_tool] 未提供数据源(--csv/--json)\n'); sys.exit(1)
    wb.save(a.output)
    print('已生成 ' + a.output + ': ' + str(len(wb.sheetnames)) + ' 个工作表(' + ', '.join(n + ' ' + str(wb[n].max_row-1) + ' 行' for n in wb.sheetnames) + ')')

def cmd_read(a):
    wb = openpyxl.load_workbook(a.input, data_only=True)
    ws = wb[a.sheet] if a.sheet else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows: print('(空表)'); return
    print('| ' + ' | '.join(str(x if x is not None else '') for x in rows[0]) + ' |')
    print('|' + '|'.join('---' for _ in rows[0]) + '|')
    for r in rows[1:]:
        print('| ' + ' | '.join(str(x if x is not None else '') for x in r) + ' |')

def cmd_tocsv(a):
    wb = openpyxl.load_workbook(a.input, data_only=True)
    ws = wb[a.sheet] if a.sheet else wb[wb.sheetnames[0]]
    with open(a.output, 'w', encoding='utf-8-sig', newline='') as f:
        w = csv.writer(f)
        for r in ws.iter_rows(values_only=True): w.writerow(['' if x is None else x for x in r])
    print('已导出 ' + a.output + ' (' + str(ws.max_row) + ' 行)')

p = argparse.ArgumentParser()
sub = p.add_subparsers(dest='cmd', required=True)
c = sub.add_parser('create'); c.add_argument('output'); c.add_argument('--csv', action='append'); c.add_argument('--json', action='append'); c.set_defaults(func=cmd_create)
r = sub.add_parser('read'); r.add_argument('input'); r.add_argument('--sheet'); r.set_defaults(func=cmd_read)
t = sub.add_parser('tocsv'); t.add_argument('input'); t.add_argument('-o','--output', required=True); t.add_argument('--sheet'); t.set_defaults(func=cmd_tocsv)
a = p.parse_args()
try:
    a.func(a)
except SystemExit:
    raise
except Exception as e:
    sys.stderr.write('[xlsx_tool] 处理错误: ' + str(e) + '\n'); sys.exit(2)
